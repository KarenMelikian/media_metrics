from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.orm import joinedload
from io import BytesIO
from openpyxl import Workbook
from fastapi import UploadFile, File
from openpyxl import load_workbook
from datetime import datetime

from core.session import SessionDep
from utils.authorization_utils import check_access_and_get_user
from schemas.user import UserSchema
from models.user_inputs import FormSubmission, FormSubmissionField

router = APIRouter()


@router.get("/export")
async def export_submissions_excel(
    session: SessionDep,
    user: UserSchema = Depends(check_access_and_get_user)
):
    stmt = (
        select(FormSubmission)
        .options(joinedload(FormSubmission.values))
        .where(FormSubmission.user_id == user.id)
    )
    result = await session.execute(stmt)
    submissions = result.unique().scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Submissions"

    headers = ["Submission ID", "Form ID", "Created At", "Field ID", "Field Value"]
    ws.append(headers)

    for sub in submissions:
        for field in sub.values:
            ws.append([
                sub.id,
                sub.form_id,
                sub.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                field.field_id,
                field.value
            ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=submissions.xlsx"}
    )



@router.post("/import")
async def import_submissions_excel(
    session: SessionDep,
    user: UserSchema = Depends(check_access_and_get_user),
    file: UploadFile = File(...)
):
    contents = await file.read()
    workbook = load_workbook(filename=BytesIO(contents))
    sheet = workbook.active

    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    current_submission = None
    current_submission_id = None

    for row in rows:
        submission_id, form_id, created_at_str, field_id, value = row

        if submission_id != current_submission_id:
            current_submission_id = submission_id
            created_at = datetime.strptime(created_at_str, "%Y-%m-%d %H:%M:%S")
            current_submission = FormSubmission(
                id=submission_id,
                form_id=form_id,
                user_id=user.id,
                created_at=created_at
            )
            session.add(current_submission)

        session.add(FormSubmissionField(
            submission_id=submission_id,
            field_id=field_id,
            value=value
        ))

    await session.commit()
    return {"detail": "Submissions imported from Excel successfully"}
